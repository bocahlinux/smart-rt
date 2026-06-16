"""Views untuk Warga (Data Warga) — Phase 3.

Lihat docs/06-API-CONTRACT.md §3 dan docs/07-TASK-BREAKDOWN.md Phase 3.
"""

import io
import uuid
from datetime import datetime

from django.http import HttpResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, permissions, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet

from accounts.models import User, WargaProfile
from accounts.permissions import IsAdmin, IsSekretaris, has_perm
from accounts.utils import error_response, success_response
from accounts.warga_filters import WargaFilter
from accounts.warga_serializers import WargaWriteSerializer, get_warga_serializer_class
from audit.services import log_action


class WargaViewSet(ModelViewSet):
    """CRUD + extra actions untuk WargaProfile.

    Endpoints:
      GET    /warga/           — list (semua role, queryset di-scope per role)
      POST   /warga/           — create (sekretaris/admin)
      GET    /warga/{id}/      — retrieve (semua role, data di-filter per role)
      PUT    /warga/{id}/      — update (sekretaris/admin)
      PATCH  /warga/{id}/      — partial update (sekretaris/admin)
      DELETE /warga/{id}/      — soft-delete (admin only)
      PUT    /warga/{id}/verify/ — verify/reject (sekretaris/admin)
      GET    /warga/export/    — export Excel/PDF (sekretaris/admin)
      POST   /warga/import/    — import Excel (sekretaris/admin)
    """

    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = WargaFilter
    ordering_fields = ["nama_lengkap", "blok", "created_at", "status"]
    ordering = ["nama_lengkap"]

    def get_queryset(self):
        user = self.request.user
        base_qs = WargaProfile.objects.select_related("user").filter(is_deleted=False)
        # Non-warga dan siapa pun yang punya izin kelola warga → lihat semua
        if user.role != "warga" or has_perm(user, "tambah_edit_warga"):
            return base_qs
        # Warga: untuk list hanya profil sendiri; untuk detail akses semua
        if self.action == "list":
            return base_qs.filter(user=user)
        return base_qs

    def get_serializer_class(self):
        if self.action in ("create", "update", "partial_update"):
            return WargaWriteSerializer
        return WargaWriteSerializer  # dipilih ulang di get_serializer per-instance

    def get_serializer(self, *args, **kwargs):
        """Pilih serializer read yang tepat berdasarkan role + target profile."""
        if self.action in ("create", "update", "partial_update"):
            kwargs["context"] = self.get_serializer_context()
            return WargaWriteSerializer(*args, **kwargs)

        instance = args[0] if args else None
        # Untuk list, instance adalah queryset; untuk retrieve, instance tunggal
        is_many = isinstance(instance, (list,)) or getattr(instance, "__iter__", None) and not isinstance(instance, WargaProfile)

        request = self.request
        user = request.user
        context = self.get_serializer_context()

        if is_many:
            from rest_framework.serializers import ListSerializer

            # Gunakan serializer admin untuk semua item bila admin/sekretaris
            SerClass = get_warga_serializer_class(user, None)
            return SerClass(instance, many=True, context=context, **{k: v for k, v in kwargs.items() if k not in ("many",)})

        target = instance
        SerClass = get_warga_serializer_class(user, target)
        return SerClass(instance, context=context, **{k: v for k, v in kwargs.items()})

    # ------------------------------------------------------------------ #
    # Permissions per action                                               #
    # ------------------------------------------------------------------ #

    def _check_write_permission(self):
        user = self.request.user
        if not has_perm(user, "tambah_edit_warga"):
            return error_response(
                "PERMISSION_DENIED",
                "Anda tidak memiliki izin untuk mengelola data warga",
                status_code=status.HTTP_403_FORBIDDEN,
            )
        return None

    def _check_admin_only(self):
        if not has_perm(self.request.user, "hapus_restore_warga"):
            return error_response(
                "PERMISSION_DENIED",
                "Anda tidak memiliki izin untuk melakukan aksi ini",
                status_code=status.HTTP_403_FORBIDDEN,
            )
        return None

    # ------------------------------------------------------------------ #
    # CRUD                                                                 #
    # ------------------------------------------------------------------ #

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)

        user = request.user
        # Pilih serializer berdasarkan role (non-owner view untuk list)
        from accounts.warga_serializers import (
            WargaAdminSerializer, WargaBendaharaSerializer,
            WargaPengurusSerializer, WargaPublicSerializer,
        )
        role_map = {
            "admin": WargaAdminSerializer,
            "sekretaris": WargaAdminSerializer,
            "bendahara": WargaBendaharaSerializer,
            "pengurus": WargaPengurusSerializer,
        }
        SerClass = role_map.get(user.role, WargaPublicSerializer)
        ctx = self.get_serializer_context()

        if page is not None:
            data = SerClass(page, many=True, context=ctx).data
            p = self.paginator
            return success_response(
                data=data,
                pagination={
                    "page": p.page.number,
                    "limit": p.page_size,
                    "total": p.page.paginator.count,
                    "totalPages": p.page.paginator.num_pages,
                },
            )

        data = SerClass(queryset, many=True, context=ctx).data
        return success_response(data=data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        user = request.user
        # Object-level: warga hanya akses profil sendiri
        if user.role == "warga" and instance.user != user:
            return error_response(
                "PERMISSION_DENIED_OBJECT_LEVEL",
                "Anda tidak memiliki akses ke data warga lain",
                status_code=status.HTTP_403_FORBIDDEN,
            )
        from accounts.warga_serializers import get_warga_serializer_class
        SerClass = get_warga_serializer_class(user, instance)
        data = SerClass(instance, context=self.get_serializer_context()).data

        log_action(
            user=user,
            action="view",
            table_name="warga_profiles",
            record_id=instance.id,
            request=request,
        )
        return success_response(data=data)

    def create(self, request, *args, **kwargs):
        denied = self._check_write_permission()
        if denied:
            return denied

        serializer = WargaWriteSerializer(data=request.data, context=self.get_serializer_context())
        if not serializer.is_valid():
            return error_response(
                "VALIDATION_ERROR",
                "Data warga tidak valid",
                errors=_serializer_errors(serializer),
                status_code=status.HTTP_400_BAD_REQUEST,
            )
        instance = serializer.save()

        log_action(
            user=request.user,
            action="create",
            table_name="warga_profiles",
            record_id=instance.id,
            new_data=_profile_dict(instance),
            request=request,
        )
        return success_response(
            data={"id": str(instance.id), "namaLengkap": instance.nama_lengkap},
            message="Data warga berhasil ditambahkan",
            status_code=status.HTTP_201_CREATED,
        )

    def update(self, request, *args, **kwargs):
        denied = self._check_write_permission()
        if denied:
            return denied

        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        old_data = _profile_dict(instance)

        serializer = WargaWriteSerializer(
            instance, data=request.data, partial=partial,
            context=self.get_serializer_context(),
        )
        if not serializer.is_valid():
            return error_response(
                "VALIDATION_ERROR",
                "Data warga tidak valid",
                errors=_serializer_errors(serializer),
                status_code=status.HTTP_400_BAD_REQUEST,
            )
        instance = serializer.save()

        log_action(
            user=request.user,
            action="update",
            table_name="warga_profiles",
            record_id=instance.id,
            old_data=old_data,
            new_data=_profile_dict(instance),
            request=request,
        )
        return success_response(
            data={"id": str(instance.id), "namaLengkap": instance.nama_lengkap},
            message="Data warga berhasil diperbarui",
        )

    def partial_update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        denied = self._check_admin_only()
        if denied:
            return denied

        instance = self.get_object()
        old_data = _profile_dict(instance)

        # Soft delete + unlink user agar user dapat dilink ke profil warga lain
        instance.is_deleted = True
        instance.deleted_at = timezone.now()
        instance.deleted_by = request.user
        instance.user = None
        instance.save(update_fields=["is_deleted", "deleted_at", "deleted_by", "user"])

        log_action(
            user=request.user,
            action="delete",
            table_name="warga_profiles",
            record_id=instance.id,
            old_data=old_data,
            request=request,
        )
        return success_response(message="Data warga berhasil dihapus")

    # ------------------------------------------------------------------ #
    # Restore / Deleted list                                               #
    # ------------------------------------------------------------------ #

    @action(detail=False, methods=["get"], url_path="deleted")
    def deleted(self, request):
        """GET /warga/deleted/ — list warga yang sudah di-soft-delete (admin only)."""
        denied = self._check_admin_only()
        if denied:
            return denied

        from accounts.warga_serializers import WargaAdminSerializer

        qs = (
            WargaProfile.objects.select_related("user", "kartu_keluarga")
            .filter(is_deleted=True)
            .order_by("-deleted_at")
        )
        data = WargaAdminSerializer(qs, many=True, context=self.get_serializer_context()).data
        return success_response(data=data)

    @action(detail=True, methods=["put"], url_path="restore")
    def restore(self, request, pk=None):
        """PUT /warga/{id}/restore/ — pulihkan warga yang di-soft-delete (admin only)."""
        denied = self._check_admin_only()
        if denied:
            return denied

        try:
            instance = WargaProfile.objects.select_related("user").get(pk=pk, is_deleted=True)
        except WargaProfile.DoesNotExist:
            return error_response(
                "NOT_FOUND",
                "Data warga tidak ditemukan atau belum dihapus",
                status_code=status.HTTP_404_NOT_FOUND,
            )

        # Cek konflik NIK dengan warga aktif lain
        if instance.nik:
            if WargaProfile.objects.filter(nik=instance.nik, is_deleted=False).exists():
                return error_response(
                    "WARGA_NIK_DUPLICATE",
                    f"NIK {instance.nik} sudah digunakan warga lain yang aktif. "
                    "Hapus atau ubah NIK warga tersebut sebelum memulihkan data ini.",
                    status_code=status.HTTP_409_CONFLICT,
                )

        # Cek apakah user sudah punya profil aktif lain
        if WargaProfile.objects.filter(user=instance.user, is_deleted=False).exclude(pk=instance.pk).exists():
            return error_response(
                "WARGA_PROFILE_EXISTS",
                "Akun user ini sudah terhubung ke profil warga lain yang aktif.",
                status_code=status.HTTP_409_CONFLICT,
            )

        instance.is_deleted = False
        instance.deleted_at = None
        instance.deleted_by = None
        instance.save(update_fields=["is_deleted", "deleted_at", "deleted_by"])

        log_action(
            user=request.user,
            action="restore",
            table_name="warga_profiles",
            record_id=instance.id,
            new_data=_profile_dict(instance),
            request=request,
        )
        return success_response(message="Data warga berhasil dipulihkan")

    @action(detail=True, methods=["post"], url_path="unlink")
    def unlink(self, request, pk=None):
        """POST /warga/{id}/unlink/ — lepas tautan user dari profil warga (admin only)."""
        denied = self._check_admin_only()
        if denied:
            return denied

        instance = self.get_object()
        if not instance.user_id:
            return error_response(
                "WARGA_NOT_LINKED",
                "Profil warga ini tidak memiliki user yang tertaut.",
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        old_user_id = instance.user_id
        instance.user = None
        instance.save(update_fields=["user"])

        log_action(
            user=request.user,
            action="unlink",
            table_name="warga_profiles",
            record_id=instance.id,
            old_data={"user_id": str(old_user_id)},
            request=request,
        )
        return success_response(message="Tautan user berhasil dilepas")

    @action(detail=True, methods=["post"], url_path="link")
    def link(self, request, pk=None):
        """POST /warga/{id}/link/ — tautkan user ke profil warga (admin only)."""
        denied = self._check_admin_only()
        if denied:
            return denied

        instance = self.get_object()
        user_id = request.data.get("userId")
        if not user_id:
            return error_response(
                "VALIDATION_ERROR",
                "userId wajib diisi",
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        try:
            target_user = User.objects.get(pk=user_id)
        except User.DoesNotExist:
            return error_response(
                "NOT_FOUND",
                "User tidak ditemukan",
                status_code=status.HTTP_404_NOT_FOUND,
            )

        # Cek apakah user sudah tertaut ke profil warga aktif lain
        existing = WargaProfile.objects.filter(user=target_user, is_deleted=False).exclude(pk=instance.pk).first()
        if existing:
            return error_response(
                "USER_ALREADY_LINKED",
                f"User ini sudah terhubung ke data warga '{existing.nama_lengkap}'. "
                "Lepas tautan terlebih dahulu sebelum menghubungkan ke profil lain.",
                status_code=status.HTTP_409_CONFLICT,
            )

        instance.user = target_user
        instance.save(update_fields=["user"])

        log_action(
            user=request.user,
            action="link",
            table_name="warga_profiles",
            record_id=instance.id,
            new_data={"user_id": str(target_user.id)},
            request=request,
        )
        return success_response(message="User berhasil ditautkan ke profil warga")

    # ------------------------------------------------------------------ #
    # Verify                                                               #
    # ------------------------------------------------------------------ #

    @action(detail=True, methods=["put"], url_path="verify")
    def verify(self, request, pk=None):
        if not has_perm(request.user, "verifikasi_warga"):
            return error_response(
                "PERMISSION_DENIED",
                "Anda tidak memiliki izin untuk memverifikasi data warga",
                status_code=status.HTTP_403_FORBIDDEN,
            )

        instance = self.get_object()
        new_status = request.data.get("status")
        if new_status not in ("active", "rejected"):
            return error_response(
                "VALIDATION_ERROR",
                "Status verifikasi harus 'active' atau 'rejected'",
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        old_data = {"status": instance.user.status}
        instance.user.status = new_status
        instance.user.save(update_fields=["status"])

        log_action(
            user=request.user,
            action="verify",
            table_name="warga_profiles",
            record_id=instance.id,
            old_data=old_data,
            new_data={"status": new_status},
            request=request,
        )
        label = "diverifikasi" if new_status == "active" else "ditolak"
        return success_response(message=f"Warga berhasil {label}")

    # ------------------------------------------------------------------ #
    # Export                                                               #
    # ------------------------------------------------------------------ #

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """GET /warga/export?fmt=excel|pdf&status=aktif&blok=A&fullData=false

        Parameter `fmt` (bukan `format`) digunakan untuk menghindari konflik
        dengan DRF URL_FORMAT_OVERRIDE yang menggunakan key 'format'.
        """
        if not has_perm(request.user, "export_import_warga"):
            return error_response(
                "PERMISSION_DENIED",
                "Anda tidak memiliki izin untuk mengekspor data warga",
                status_code=status.HTTP_403_FORBIDDEN,
            )

        fmt = request.query_params.get("fmt", "excel").lower()
        if fmt not in ("excel", "pdf"):
            return error_response(
                "VALIDATION_ERROR",
                "Parameter 'format' harus 'excel' atau 'pdf'",
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        full_data = request.query_params.get("fullData", "false").lower() == "true"
        # fullData=true hanya untuk admin — lihat docs/06-API-CONTRACT.md §3.7
        if full_data and not has_perm(request.user, "hapus_restore_warga"):
            full_data = False

        qs = self.filter_queryset(self.get_queryset())

        log_action(
            user=request.user,
            action="export",
            table_name="warga_profiles",
            record_id=uuid.uuid4(),
            new_data={"fmt": fmt, "count": qs.count(), "fullData": full_data},
            request=request,
        )

        if fmt == "excel":
            return self._export_excel(qs, full_data)
        return self._export_pdf(qs, full_data)

    def _export_excel(self, qs, full_data: bool):
        try:
            import openpyxl
        except ImportError:
            return error_response(
                "VALIDATION_ERROR",
                "Library openpyxl tidak tersedia di server",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        from accounts.warga_serializers import _mask_nik, _mask_phone

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Warga"

        if full_data:
            headers = ["No", "NIK", "Nama Lengkap", "Blok", "No Rumah",
                       "No KK", "Hubungan KK", "Tempat Lahir", "Tanggal Lahir",
                       "Jenis Kelamin", "Agama", "Status Perkawinan",
                       "Pendidikan", "Pekerjaan", "Alamat", "No HP", "Email", "Status"]
        else:
            headers = ["No", "NIK (Masked)", "Nama Lengkap", "Blok", "No Rumah",
                       "Status Perkawinan", "Pekerjaan", "Status"]

        ws.append(headers)

        for i, warga in enumerate(qs, start=1):
            if full_data:
                ws.append([
                    i,
                    warga.nik or "",
                    warga.nama_lengkap,
                    warga.blok or "",
                    warga.no_rumah or "",
                    warga.kartu_keluarga.no_kk if warga.kartu_keluarga else "",
                    warga.get_hubungan_keluarga_display() if warga.hubungan_keluarga else "",
                    warga.tempat_lahir or "",
                    str(warga.tanggal_lahir) if warga.tanggal_lahir else "",
                    warga.get_jenis_kelamin_display() if warga.jenis_kelamin else "",
                    warga.agama or "",
                    warga.get_status_perkawinan_display() if warga.status_perkawinan else "",
                    warga.pendidikan or "",
                    warga.pekerjaan or "",
                    warga.alamat or "",
                    warga.user.phone or "",
                    warga.user.email or "",
                    warga.status,
                ])
            else:
                ws.append([
                    i,
                    _mask_nik(warga.nik),
                    warga.nama_lengkap,
                    warga.blok or "",
                    warga.no_rumah or "",
                    warga.get_status_perkawinan_display() if warga.status_perkawinan else "",
                    warga.pekerjaan or "",
                    warga.status,
                ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"data-warga-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _export_pdf(self, qs, full_data: bool):
        try:
            from weasyprint import HTML
        except ImportError:
            return error_response(
                "VALIDATION_ERROR",
                "Library WeasyPrint tidak tersedia di server",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        from accounts.warga_serializers import _mask_nik

        rows_html = ""
        for i, w in enumerate(qs, start=1):
            nik = w.nik if full_data else _mask_nik(w.nik)
            rows_html += (
                f"<tr><td>{i}</td><td>{nik or '-'}</td>"
                f"<td>{w.nama_lengkap}</td><td>{w.blok or '-'}</td>"
                f"<td>{w.no_rumah or '-'}</td><td>{w.status}</td></tr>"
            )

        html_content = f"""
        <!DOCTYPE html><html><head>
        <meta charset="utf-8">
        <style>body{{font-family:sans-serif;font-size:11px;}}
        table{{border-collapse:collapse;width:100%;}}
        th,td{{border:1px solid #ccc;padding:4px 8px;}}
        th{{background:#f0f0f0;}}</style>
        </head><body>
        <h2>Data Warga RT</h2>
        <p>Diekspor: {datetime.now().strftime('%d %B %Y %H:%M')}</p>
        <table><thead><tr>
        <th>No</th><th>NIK</th><th>Nama Lengkap</th><th>Blok</th><th>No Rumah</th><th>Status</th>
        </tr></thead><tbody>{rows_html}</tbody></table>
        </body></html>"""

        pdf = HTML(string=html_content).write_pdf()
        filename = f"data-warga-{datetime.now().strftime('%Y%m%d-%H%M%S')}.pdf"
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    # ------------------------------------------------------------------ #
    # Import                                                               #
    # ------------------------------------------------------------------ #

    @action(detail=False, methods=["post"], url_path="import", parser_classes=[MultiPartParser])
    def import_excel(self, request):
        """POST /warga/import — import data warga dari file Excel."""
        if not has_perm(request.user, "export_import_warga"):
            return error_response(
                "PERMISSION_DENIED",
                "Anda tidak memiliki izin untuk mengimpor data warga",
                status_code=status.HTTP_403_FORBIDDEN,
            )

        file = request.FILES.get("file")
        if not file:
            return error_response(
                "VALIDATION_ERROR",
                "File wajib diunggah",
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        allowed_types = {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        }
        if file.content_type not in allowed_types and not file.name.endswith((".xlsx", ".xls")):
            return error_response(
                "FILE_TYPE_NOT_ALLOWED",
                "Hanya file Excel (.xlsx / .xls) yang diizinkan",
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            )

        if file.size > 5 * 1024 * 1024:
            return error_response(
                "FILE_TOO_LARGE",
                "Ukuran file melebihi batas maksimum 5MB",
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            )

        try:
            import openpyxl
        except ImportError:
            return error_response(
                "VALIDATION_ERROR",
                "Library openpyxl tidak tersedia di server",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception:
            return error_response(
                "WARGA_IMPORT_FORMAT_INVALID",
                "File tidak dapat dibaca. Pastikan format file Excel valid.",
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            )

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        imported = 0
        failed = 0
        errors = []

        # Kolom yang diharapkan (sesuai template export):
        # 0:No, 1:NIK, 2:Nama Lengkap, 3:Blok, 4:No Rumah,
        # 5:No KK, 6:Hubungan KK, 7:Tempat Lahir, 8:Tanggal Lahir,
        # 9:Jenis Kelamin, 10:Agama, 11:Status Perkawinan, 12:Pendidikan,
        # 13:Pekerjaan, 14:Alamat, 15:No HP, 16:Email, 17:Status
        for row_num, row in enumerate(rows, start=2):
            try:
                if not row or not row[2]:
                    continue

                nik = str(row[1]).strip() if row[1] else None
                nama_lengkap = str(row[2]).strip()
                blok = str(row[3]).strip() if row[3] else None
                no_rumah = str(row[4]).strip() if row[4] else None
                no_kk_raw = str(row[5]).strip() if row[5] else None
                hubungan_keluarga_raw = str(row[6]).strip() if row[6] else None
                tempat_lahir = str(row[7]).strip() if row[7] else None
                tanggal_lahir_raw = row[8]
                jenis_kelamin = str(row[9]).strip()[:1].upper() if row[9] else None
                agama = str(row[10]).strip() if row[10] else None
                status_perkawinan_raw = str(row[11]).strip() if row[11] else None
                pendidikan = str(row[12]).strip() if row[12] else None
                pekerjaan = str(row[13]).strip() if row[13] else None
                alamat = str(row[14]).strip() if row[14] else None
                phone = str(row[15]).strip() if row[15] else None
                email = str(row[16]).strip() if row[16] else None
                warga_status = str(row[17]).strip().lower() if len(row) > 17 and row[17] else "aktif"

                if nik and WargaProfile.objects.filter(nik=nik, is_deleted=False).exists():
                    errors.append(f"Baris {row_num}: NIK {nik} sudah terdaftar")
                    failed += 1
                    continue

                # Parse tanggal lahir
                tanggal_lahir = None
                if tanggal_lahir_raw:
                    if isinstance(tanggal_lahir_raw, datetime):
                        tanggal_lahir = tanggal_lahir_raw.date()
                    else:
                        try:
                            from datetime import date
                            tanggal_lahir = date.fromisoformat(str(tanggal_lahir_raw))
                        except (ValueError, TypeError):
                            pass

                # Normalize status_perkawinan
                sp_map = {
                    "belum kawin": "belum_kawin",
                    "kawin": "kawin",
                    "cerai hidup": "cerai_hidup",
                    "cerai mati": "cerai_mati",
                }
                status_perkawinan = sp_map.get(
                    (status_perkawinan_raw or "").lower().strip(), None
                )

                # Hubungkan ke KartuKeluarga jika no_kk ada
                kk_instance = None
                if no_kk_raw:
                    from kartu_keluarga.models import KartuKeluarga  # noqa: PLC0415
                    kk_instance, _ = KartuKeluarga.objects.get_or_create(
                        no_kk=no_kk_raw,
                        defaults={"alamat": alamat or "", "created_by": request.user},
                    )

                # Map hubungan_keluarga ke choice value
                hub_map = {
                    "kepala keluarga": "kepala_keluarga",
                    "istri": "istri",
                    "anak": "anak",
                    "orang tua": "orang_tua",
                    "menantu": "menantu",
                    "cucu": "cucu",
                    "saudara": "saudara",
                }
                hubungan_keluarga = hub_map.get(
                    (hubungan_keluarga_raw or "").lower().strip(), "lainnya"
                ) if hubungan_keluarga_raw else None

                WargaProfile.objects.create(
                    user=None,
                    nik=nik,
                    nama_lengkap=nama_lengkap,
                    tempat_lahir=tempat_lahir,
                    tanggal_lahir=tanggal_lahir,
                    jenis_kelamin=jenis_kelamin if jenis_kelamin in ("L", "P") else None,
                    agama=agama,
                    status_perkawinan=status_perkawinan,
                    pendidikan=pendidikan,
                    pekerjaan=pekerjaan,
                    kartu_keluarga=kk_instance,
                    hubungan_keluarga=hubungan_keluarga,
                    alamat=alamat,
                    blok=blok,
                    no_rumah=no_rumah,
                    status=warga_status if warga_status in ("aktif", "tidak_aktif", "pindah", "meninggal") else "aktif",
                )
                imported += 1

            except Exception as exc:
                errors.append(f"Baris {row_num}: {str(exc)}")
                failed += 1

        log_action(
            user=request.user,
            action="import",
            table_name="warga_profiles",
            record_id=uuid.uuid4(),
            new_data={"imported": imported, "failed": failed},
            request=request,
        )

        return success_response(
            data={"imported": imported, "failed": failed, "errors": errors[:20]},
            message=f"Import selesai: {imported} berhasil, {failed} gagal",
        )


# ------------------------------------------------------------------ #
# Helpers                                                              #
# ------------------------------------------------------------------ #

class ProfilSayaView(APIView):
    """GET/POST /warga/profil-saya/ — buat atau cek profil sendiri tanpa perlu izin tambah_edit_warga.

    Berlaku untuk semua role yang sudah login (ketua_rt, pengurus, bendahara, sekretaris, warga).
    """

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            profile = WargaProfile.objects.get(user=request.user, is_deleted=False)
            return success_response({
                "id": str(profile.id),
                "namaLengkap": profile.nama_lengkap,
                "nik": profile.nik,
                "hasProfile": True,
            })
        except WargaProfile.DoesNotExist:
            return success_response({"hasProfile": False})

    def post(self, request):
        if WargaProfile.objects.filter(user=request.user, is_deleted=False).exists():
            return error_response(
                "PROFILE_EXISTS",
                "Anda sudah memiliki profil warga.",
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        serializer = WargaWriteSerializer(data=request.data, context={"request": request})
        if not serializer.is_valid():
            return error_response(
                "VALIDATION_ERROR",
                "Data tidak valid",
                errors=_serializer_errors(serializer),
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        vd = dict(serializer.validated_data)
        vd.pop("userId", None)
        instance = WargaProfile.objects.create(user=request.user, **vd)
        log_action(
            user=request.user,
            action="create",
            table_name="warga_profiles",
            record_id=instance.id,
            new_data=_profile_dict(instance),
            request=request,
        )
        return success_response(
            data={"id": str(instance.id), "namaLengkap": instance.nama_lengkap},
            message="Profil berhasil dibuat.",
            status_code=status.HTTP_201_CREATED,
        )


def _profile_dict(profile: WargaProfile) -> dict:
    """Ambil dict data profil untuk audit log (field sensitif akan di-mask oleh audit.services)."""
    return {
        "nik": profile.nik,
        "nama_lengkap": profile.nama_lengkap,
        "tempat_lahir": profile.tempat_lahir,
        "tanggal_lahir": str(profile.tanggal_lahir) if profile.tanggal_lahir else None,
        "blok": profile.blok,
        "no_rumah": profile.no_rumah,
        "status": profile.status,
        "no_kk": profile.kartu_keluarga.no_kk if profile.kartu_keluarga else None,
        "alamat": profile.alamat,
        "phone": getattr(profile.user, "phone", None),
        "email": getattr(profile.user, "email", None),
    }


def _serializer_errors(serializer) -> list:
    errors = []
    for field, messages in serializer.errors.items():
        for msg in messages:
            errors.append({"field": field, "message": str(msg)})
    return errors
